# -*- coding: UTF-8 ...
import re
import requests
from bs4 import BeautifulSoup
from  fake_useragent import UserAgent
import os
import telegram
import smtplib
from email.mime.text import MIMEText
from email.header import Header
import time
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
import openpyxl
import random




def get_xila_ip(url,new_ip):
    '''
    从免费代理IP网站获取代理IP
    '''

    # url = 'http://www.xiladaili.com/gaoni/'

    headers= {
        # 'User-Agent': UserAgent().random
        'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding':'gzip, deflate',
        'Accept-Language':'zh-CN,zh;q=0.9,zh-TW;q=0.8,en;q=0.7',
        'Cache-Control':'max-age=0',
        'Connection':'keep-alive',
        'DNT': '1',
        'Host': 'www.xiladaili.com',
        # 'Referer':'http://www.xiladaili.com/gaoni/7/',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.101 Safari/537.36',

    }

    proxy = {'http': new_ip, 'https': new_ip}

    respons = requests.get(url=url,headers=headers,proxies= proxy)

    # print(respons.status_code)
    # print(respons.text)

    bs =BeautifulSoup(respons.content,'html.parser')
    info = bs.find('tbody')
    # print(info)
    data = info.find_all('tr')
    # print(data)

    ip_list=[]   #将获取到的IP用列表存储
    for i in data:
        dic = {}
        ip_port = i.find_all('td')[0].string          #拆分IP和端口
        tpy = i.find_all('td')[1].string              #拆分类型 http或https
        time_yanchi = i.find_all('td')[4].string      #响应时间
        core = i.find_all('td')[7].string             #打分


        # if tpy != 'HTTP代理':           #把https存起来，因为使用中主要还是使用https
        dic['ip_port']=ip_port
        dic['tpy']=tpy
        dic['time_yanchi']=time_yanchi
        dic['core']=core

        ip_list.append(dic)    #用字典形式存入列表中

        # print(f'ip和端口：{ip_port},类型：{tpy},响应时间：{time_yanchi},分数：{core}')
    print(ip_list)
    if len(ip_list) != 0:
        print('xiladaili爬取成功')
    else:
        print('xiladaili爬取失败')

    return ip_list   #将获取到的IP列表返回


def get_kuaidaili_ip(url):
    '''
    从免费代理IP网站获取代理IP
    '''

    # url = 'http://www.xiladaili.com/gaoni/'

    headers= {
        # 'User-Agent': UserAgent().random
        'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding':'gzip,deflate,br',
        'Accept-Language':'zh-CN,zh;q=0.9,zh-TW;q=0.8,en;q=0.7',
        'Cache-Control':'max-age=0',
        'Connection':'keep-alive',
        'cookie':'channelid=0; sid=1623726223825649; _ga=GA1.2.1408876963.1623727124; _gcl_au=1.1.573137817.1624064417; _gid=GA1.2.1635708697.1624064417; Hm_lvt_7ed65b1cc4b810e9fd37959c9bb51b31=1624064417; Hm_lpvt_7ed65b1cc4b810e9fd37959c9bb51b31=1624064417',
        'DNT': '1',
        'Host': 'www.kuaidaili.com',
        'Referer':'https://www.google.com/',
        'sec-ch-ua':'" Not;A Brand";v="99", "Google Chrome";v="91", "Chromium";v="91"',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.101 Safari/537.36',

    }

    # proxy = {'http': new_ip, 'https': new_ip}

    respons = requests.get(url=url,headers=headers)#,proxies= proxy)

    # print(respons.status_code)
    # print(respons.text)

    bs =BeautifulSoup(respons.content,'html.parser')
    info = bs.find('tbody')
    # print(info)
    data = info.find_all('tr')
    # print(data)

    ip_list=[]   #将获取到的IP用列表存储
    for i in data:
        dic = {}
        ip = i.find_all('td')[0].string          #拆分IP和端口
        port = i.find_all('td')[1].string          #拆分IP和端口
        tpy = i.find_all('td')[3].string              #拆分类型 http或https
        time_yanchi = i.find_all('td')[5].string      #响应时间
        core = i.find_all('td')[6].string             #打分


        # if tpy != 'HTTP代理':           #把https存起来，因为使用中主要还是使用https
        dic['ip_port']=ip+':'+port
        dic['tpy']=tpy
        dic['time_yanchi']=time_yanchi
        dic['core']=core

        ip_list.append(dic)    #用字典形式存入列表中

        # print(f'ip和端口：{ip_port},类型：{tpy},响应时间：{time_yanchi},分数：{core}')
    print(ip_list)
    if len(ip_list) != 0:
        print('kuaidaili爬取成功')
    else:
        print('kuaidaili爬取失败')
    return ip_list   #将获取到的IP列表返回



# get_ip('http://www.xiladaili.com/gaoni/6/','')

def check_if_ip_in_excel(ip_list):
    # 先读取，再写入
    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\代理IP.xlsx')
    ws = book.active  # 获取当前正在操作的表对象
    strs = []
    for key_word in ws['A']:
        strs.append(key_word.value)

    useful_iplist=[]
    for ip in ip_list:
        if ip['ip_port'] not in strs:
            useful_iplist.append(ip)
            # return ip
        else:
            # return
            print(f'{ip}已经存在，不检查')

    return useful_iplist

def check_ip(ip_list,num):
    '''
    检测IP是否有效，但在pycharm中一直返回的是本身IP，而在vps中运行时则会正常。。。不知道怎么回事
    '''
    # 访问网址
    # url = 'http://www.whatismyip.com.tw/'
    url = 'http://icanhazip.com'  # 用来检测IP是否可以正常使用
    # url = 'http://httpbin.org/ip'

    # 这是代理IP
    ip_useful_list = []
    for dic_ip in ip_list:
        ip = dic_ip['ip_port']
        # 设置代理ip访问方式，http和https
        proxy = {'http': ip, 'https': ip}
        # 创建ProxyHandler
        try:
            response = requests.get(url=url, proxies=proxy, timeout=2)

            # print(response.status_code)
            # print(response.text)
            print(f'{dic_ip}有效')
            dic_ip['useful'] = 111
            num += 1

        except:
            print(f'{dic_ip}无效')
            dic_ip['useful'] = 000
            # continue
            num = num

        ip_useful_list.append(dic_ip)  # 把IP添加到列表中,是否有效通过useful来标记
    return ip_useful_list,num


def write_excel(list):
    """将传入的数据写入"""
    #先读取，再写入
    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\代理IP.xlsx')
    ws = book.active  # 获取当前正在操作的表对象

    #把ip提取出来,判断是否存在

    # strs = []
    # for key_word in ws['A']:
    #     strs.append(key_word.value)

    j = ws.max_row  #获取行数，添加到后面去，就不会把前面的覆盖掉
    for ip_info in list:
        # ip_port = ip_info['ip_port']
        # if ip_port not in strs:
        # print(f'{ip_info}  不存在,存入')
            # text = list[0] + '  ' + list[1] + '  ' + list[2] + '\n'
            # send_telegram(text)    #关键词之前不存在,则推送并添加进去
            # ws.append(ip_info)   #不能直接写进去，要先读取字典

        j += 1
            # for item in ip_info.items():
            # j +=1
        ws['A'+str(j)].value=ip_info['ip_port']
        ws['B'+str(j)].value=ip_info['tpy']
        ws['C'+str(j)].value=ip_info['time_yanchi']
        ws['D'+str(j)].value=ip_info['core']
        ws['E'+str(j)].value=ip_info['useful']


        book.save("C:\\Users\\Administrator\\Desktop\\代理IP.xlsx")
            # print(f'{ip_port}写入成功')




def get_excel_ip():
    '''
    从excle文件中随机读取一个ip，然后判断是否有效，并输出
    :return:
    '''

    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\代理IP.xlsx')
    ws = book.active  # 获取当前正在操作的表对象

    k = ws.max_row
    while True:
        i = random.randint(1,k)     #获取随机一行IP
        if ws['E'+str(i)].value == 111:       #先通过useful这个值来判断是否有效，有效则再来检测
            ip_port = ws['A'+str(i)].value
            # print(ip_port)

            url = 'http://icanhazip.com'  # 用来检测IP是否可以正常使用
            proxy = {'http':ip_port, 'https':ip_port}
            try:
                response = requests.get(url=url, proxies=proxy, timeout=2)

                # print(f'{ip_port}有效')
                return ip_port
            except:
                # print(f'{ip_port}无效')
                ws['E' + str(i)].value = 000
                # ws.delete_rows(i)  # 无效则删除第i行，后面数据补充上去    #不删除
                book.save("C:\\Users\\Administrator\\Desktop\\代理IP.xlsx")



def delate_ip(ip):

    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\代理IP.xlsx')
    ws = book.active  # 获取当前正在操作的表对象

    strs = []
    for key_word in ws['A']:
        strs.append(key_word.value)
    if ip in strs:
        j = strs.index(ip)+1
        ws['E' + str(j)].value = 000
        # ws.delete_rows(j)  # 无效则删除第i行，后面数据补充上去
        book.save("C:\\Users\\Administrator\\Desktop\\代理IP.xlsx")
        print(f'{ip}无效,重新标记')


def send_wechat(title,text,detal_url):

    get_acs_token_url = 'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid=ww428187fbdb9fd50f&corpsecret=po3lAng2YdZzaa67teFhmOxL34EuuWpffdiZlQeGaNk'

    access = requests.get(url=get_acs_token_url)
    access_token = eval(access.text)['access_token']
    # print(type(eval(access_token.text)))
    # print(access_token)


    url = f'https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token={access_token}'
    # print(url)

    headers = {
        'User-Agent': UserAgent().random
    }

    key = {
        "touser": "@all",
        # "toparty": "PartyID1|PartyID2",
        # "totag": "TagID1 | TagID2",
        "msgtype": "textcard",
        "agentid": 1000002,
        # "text": '测试',
        "textcard": {
            "title": title,
            "description": text,#"<div class=\"gray\">2016年9月26日</div> <div class=\"normal\">恭喜你抽中iPhone 7一台，领奖码：xxxx</div><div class=\"highlight\">请于2016年10月10日前联系行政同事领取</div>",
            "url": detal_url,
            "btntxt": "更多"
        },
        "safe": 0,
        "enable_id_trans": 0,
        "enable_duplicate_check": 0,
        "duplicate_check_interval": 1800
    }

    res = requests.post(url=url,headers=headers,json=key)
    # print(res.status_code)
    # print(res.text)

# send_wechat('测试','你好啊','URL')


def send_telegram(text):
    chat_id = '1203976293'
    token = '1805409515:AAEQDXHNpZkgRe27udY99D5Rv3q9cNhDpyI'
    bot = telegram.Bot(token=token)

    bot.send_message(chat_id=chat_id, text=text)

def get_time():
    # 格式化成2016-03-20 11:45:39形式
    real_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    # print (real_time)
    return real_time

def main():

    new_ip = get_excel_ip()
    i = 1
    j = 1

    start_time = time.time()
    num = 0
    while i <=5:
    # for i in range(1,25):   #把前四页的爬取出来

        print(new_ip)
        try:

            url1 = 'http://www.xiladaili.com/gaoni/'+str(i)+'/'
            url2 = 'https://www.kuaidaili.com/free/intr/'+str(i)+'/'
            ip_list1 = get_xila_ip(url1,new_ip)   #先爬取IP
            ip_list2 = get_kuaidaili_ip(url2)   #先爬取IP
            # print(ip_list)

            # ip_list = ip_list2
            ip_list = ip_list1 + ip_list2
            # print(ip_list)


            # print(ip_useful_list)
            if len(ip_list) == 0:    #判断是否爬取到了

                delate_ip(new_ip)
                new_ip = get_excel_ip()
                print(f'第{i}页ip爬取出错{j}次，更换IP:{new_ip}重新爬取')
                if j >20:
                    print(f'出错20次，跳过第{i}页')
                    i+=1
                    j=0
                j += 1
                time.sleep(5)
                num = num
            else:
                ip_notin_excel_list = check_if_ip_in_excel(ip_list)   #先判断IP是否存在于excel中，把不存在的输出来
                get_touple = check_ip(ip_notin_excel_list,num)    #然后判断是否有效
                ip_useful_list = get_touple[0]     #然后判断是否有效
                num_old = get_touple[1]

                # print(ip_useful_list)
                write_excel(ip_useful_list)

                print(f'第{i}页ip爬取写入成功')
                i += 1
                j = 1
                num = num_old



        except:
            delate_ip(new_ip)
            new_ip = get_excel_ip()
            print(f'第{i}页ip爬取出错{j}次，更换IP:{new_ip}重新爬取')
            if j >20:
                print(f'出错20次，跳过第{i}页')
                i+=1
                j=0
            j+=1
            time.sleep(5)
            num = num



    #添加计算采集时间功能
    end_time = time.time()
    seconds = int(end_time - start_time)
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    time_used = "%d:%02d:%02d" % (h, m, s)

    # print("%d:%02d:%02d" % (h, m, s))

    send_text = '爬取完成,用时'+str(time_used)+',共采集了'+str(num)+'个有效IP'

    print(send_text)
    send_wechat('IP池爬取', send_text, 'URL')
    send_telegram(send_text)

    #从exel中获取一个有效ip并返回
    # ip_useful = get_excel()
    # print(ip_useful)


main()
