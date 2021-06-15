# -*- coding: UTF-8 ...
import re
import requests
from bs4 import BeautifulSoup
from  fake_useragent import UserAgent
import os
import telegram
import smtplib
from email.mime.text import MIMEText
from email.header import Header
import time
import random
import openpyxl




def get_time():
    # 格式化成2016-03-20 11:45:39形式
    real_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    # print (real_time)
    return real_time

def send_telegram(text):
    chat_id = '1203976293'
    token = '1886850695:AAG-gfzszOsF_RPJHdYCaI0maxszwd4qoL4'
    bot = telegram.Bot(token=token)

    bot.send_message(chat_id=chat_id, text=text)


def send_wechat(title,text,detal_url):

    get_acs_token_url = 'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid=ww428187fbdb9fd50f&corpsecret=po3lAng2YdZzaa67teFhmOxL34EuuWpffdiZlQeGaNk'

    access = requests.get(url=get_acs_token_url)
    access_token = eval(access.text)['access_token']
    # print(type(eval(access_token.text)))
    # print(access_token)


    url = f'https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token={access_token}'
    # print(url)

    headers = {
        'User-Agent': UserAgent().random
    }

    key = {
        "touser": "@all",
        # "toparty": "PartyID1|PartyID2",
        # "totag": "TagID1 | TagID2",
        "msgtype": "textcard",
        "agentid": 1000002,
        # "text": '测试',
        "textcard": {
            "title": title,
            "description": text,#"<div class=\"gray\">2016年9月26日</div> <div class=\"normal\">恭喜你抽中iPhone 7一台，领奖码：xxxx</div><div class=\"highlight\">请于2016年10月10日前联系行政同事领取</div>",
            "url": detal_url,
            "btntxt": "更多"
        },
        "safe": 0,
        "enable_id_trans": 0,
        "enable_duplicate_check": 0,
        "duplicate_check_interval": 1800
    }

    res = requests.post(url=url,headers=headers,json=key)
    # print(res.status_code)
    # print(res.text)

# send_wechat('测试','你好啊','URL')


'''

def check_ip(ip_list):
    #访问网址
    # url = 'http://www.whatismyip.com.tw/'
    url = 'http://icanhazip.com'
    # url = 'http://httpbin.org/ip'
    #这是代理IP
    # ip = '125.177.124.73:80'

    ip_useful_list = []

    for ip in ip_list:
        # ip = '103.103.3.6:8080'
        # ip = '131.161.68.37:31264'
        #设置代理ip访问方式，http和https
        proxy = {'http':ip,'https':ip}
        # proxy = {'http':'http://103.103.3.6:8080','https':'https://103.103.3.6:8080'}
        #创建ProxyHandler
        try:
            response = requests.get(url=url,proxies=proxy,timeout=2)

            # print(response.status_code)
            # print(response.text)
            print(f'{ip}有效')
            ip_useful_list.append(ip)


        except:
            print(f'{ip}无效')

    return ip_useful_list

def get_ip():

    url = 'http://www.xiladaili.com/gaoni/'

    headers= {
        'User-Agent': UserAgent().random
    }

    respons = requests.get(url=url,headers=headers)

    bs =BeautifulSoup(respons.content,'html.parser')
    info = bs.find('tbody')
    data = info.find_all('tr')
    # print(data)
    ip_list=[]
    for i in data:
        # print(str(i))
        # detal = str(i)
        ip_port = i.find_all('td')[0].string
        tpy = i.find_all('td')[1].string
        time_yanchi = i.find_all('td')[4].string
        core = i.find_all('td')[7].string
        if tpy != 'HTTP代理':
            ip_list.append(ip_port)
        # ip_list.append(ip_port)

        # print(f'ip和端口：{ip_port},类型：{tpy},响应时间：{time_yanchi},分数：{core}')

        # for j in i.find_all('td'):
        #     print(j)
        # print(type(detal))

    # print(type(data))
    # print(ip_list)



    # print(respons.status_code)
    # print(respons.text)

    return ip_list
'''


def get_excel():
    '''
    从excle文件中随机读取一个ip，然后判断是否有效，并输出
    :return:
    '''

    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\代理IP.xlsx')
    ws = book.active  # 获取当前正在操作的表对象

    k = ws.max_row
    while True:
        i = random.randint(2,k)   #获取随机的一个IP
        ip_port = ws['A'+str(i)].value
        # print(ip_port)

        url = 'http://icanhazip.com'  # 用来检测IP是否可以正常使用
        proxy = {'http':ip_port, 'https':ip_port}
        try:
            response = requests.get(url=url, proxies=proxy, timeout=2)

            # print(f'{ip_port}有效')
            return ip_port
        except:
            # print(f'{ip_port}无效')
            ws.delete_rows(i)  # 无效则删除第i行，后面数据补充上去
            book.save("C:\\Users\\Administrator\\Desktop\\代理IP.xlsx")


def delate_ip(ip):

    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\代理IP.xlsx')
    ws = book.active  # 获取当前正在操作的表对象

    strs = []
    for key_word in ws['A']:
        strs.append(key_word.value)

    j = strs.index(ip)+1   #要加1才是真实行
    ws.delete_rows(j)  # 无效则删除第i行，后面数据补充上去
    book.save("C:\\Users\\Administrator\\Desktop\\代理IP.xlsx")
    print(f'{ip}无效，删除')



def get_in(ip):
    proxy = {'https': ip}
    print(ip)
    url = 'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=auth&code=073YxKkl2usAa74gpcnl2ay1sO1YxKkH&key302=2caf893321&expire302=1623027326'

    headers = {
        'Host': 'cloud.cn2030.com',
        'Connection': 'keep - alive',
        'Cookie':'',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36 MicroMessenger/7.0.9.501 NetType/WIFI MiniProgramEnv/Windows WindowsWechat',
        'content-type': 'application/json',
        'zftsl': 'd3916c78ae31015310937eead3d83093',
        'Referer': 'https://servicewechat.com/wx2c7f0f3c30d99445/73/page-frame.html',
        'Accept-Encoding': 'gzip, deflate, br',

    }

    response = requests.get(url=url,headers=headers,proxies=proxy)

    print(response.status_code)
    # print(response.headers)
    info = eval(str(response.headers))
    print(info)
    set_cookie = info['Set-Cookie'].split(';',2)[0]
    print(set_cookie)
    return set_cookie


# get_in()

def get_info(url,cookie,k,ip):
    proxy = {'https': ip}
    # url = 'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=3653&lat=30.27415&lng=120.15515&key302=06da522bf2&expire302=1623028311'
    # url = 'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4846&lat=30.27415&lng=120.15515&key302=2b7b75179a&expire302=1623027327'
    # url = 'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4851&lat=30.27415&lng=120.15515&key302=ae9b132ff0&expire302=1623030780'
    # url = 'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4851&lat=30.27415&lng=120.15515&key302=ae9b132ff0&expire302=1623030780'

    headers = {
        'Host': 'cloud.cn2030.com',
        'Connection': 'keep - alive',
        'Cookie': cookie,
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36 MicroMessenger/7.0.9.501 NetType/WIFI MiniProgramEnv/Windows WindowsWechat',
        'content-type': 'application/json',
        'zftsl': 'cf6c3f398e93c4b98f9b43dbe3c2814f',
        'Referer': 'https://servicewechat.com/wx2c7f0f3c30d99445/73/page-frame.html',
        'Accept-Encoding': 'gzip, deflate, br',

    }

    response = requests.get(url=url,headers=headers,proxies=proxy)
    # print(response.status_code)
    # print(response.text)
    info = eval(str(response.text.replace('true','111').replace('false','000')))
    # print(type(info))
    addr = info['cname']+':'+info['addr']

    data = eval(str(info['list']))

    # print(f'地址：{addr}')
    # print(type(data))
    # print(len(data))
    text = f'地址：{addr}' + '\n'
    for d in data:
        i = eval(str(d))
        # print(f'{i["text"]} 价格:{i["price"]}   {i["BtnLable"]}， 预约时间：{i["date"]}')
        if str(i["date"])  != '暂无' :
            send_telegram(f'{i["text"]} 价格:{i["price"]}   {i["BtnLable"]}， 预约时间：{i["date"]}  地址：{addr}，请登录知苗预约小程序准备预约')
            send_wechat('九价疫苗',f'{i["text"]} 价格:{i["price"]}   {i["BtnLable"]}， 预约时间：{i["date"]}  地址：{addr}，请登录知苗预约小程序准备预约',f'{url}')
        elif str(i["BtnLable"]) != '暂未开始':
            send_telegram(f'{i["text"]} 价格:{i["price"]}   {i["BtnLable"]}， 预约时间：{i["date"]}  地址：{addr}，请登录知苗预约小程序准备预约')
            send_wechat('九价疫苗',f'{i["text"]} 价格:{i["price"]}   {i["BtnLable"]}， 预约时间：{i["date"]}  地址：{addr}，请登录知苗预约小程序准备预约',f'{url}')
        else:
            k += 1

        text = text + f'{i["text"]} 价格:{i["price"]}   {i["BtnLable"]}， 预约时间：{i["date"]}' + '\n'
    # print(text)
    return text,k

        # print(type(eval(str(d))))
        # print(d)


    # print(data[0])
    # print(data[1])


# get_info()

def main():
    # ip_list = get_ip()
    #
    # ip_useful_list = check_ip(ip_list)
    #
    # print(ip_useful_list)

    ip = get_excel()
    # cookie = get_in(ip)
    m = 0
    while True:

        try:
            cookie = get_in(ip)
            break
        except:
            time.sleep(2)
            m +=1
            print(f'获取cookie失败{m}次')
            if m >5:
                delate_ip(ip)
                ip = get_excel()
                print(f'获取cookie5次失败，更换IP：{ip}')
                m = 0


    url_list = ['https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4846&lat=30.27415&lng=120.15515&key302=2b7b75179a&expire302=1623027327',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=3653&lat=30.27415&lng=120.15515&key302=06da522bf2&expire302=1623028311',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4851&lat=30.27415&lng=120.15515&key302=ae9b132ff0&expire302=1623030780',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4231&lat=30.27415&lng=120.15515&key302=f49f06d8a7&expire302=1623030939',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4852&lat=30.27415&lng=120.15515&key302=174c49fc14&expire302=1623030980',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=1755&lat=30.27415&lng=120.15515&key302=d6901e7a0c&expire302=1623030991',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=1756&lat=30.27415&lng=120.15515&key302=303fc255fc&expire302=1623031022',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4460&lat=30.27415&lng=120.15515&key302=8660f5f3df&expire302=1623031036',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4220&lat=30.27415&lng=120.15515&key302=3b8d039a00&expire302=1623031046',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4475&lat=30.27415&lng=120.15515&key302=27e8a390b6&expire302=1623031057',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4845&lat=30.27415&lng=120.15515&key302=88639f2395&expire302=1623031062',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4849&lat=30.27415&lng=120.15515&key302=ec6d00df21&expire302=1623031077',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4850&lat=30.27415&lng=120.15515&key302=b525ea43e4&expire302=1623031085',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4848&lat=30.27415&lng=120.15515&key302=eaa22b8a8f&expire302=1623031111',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=5172&lat=30.27415&lng=120.15515&key302=eb7d0f49c4&expire302=1623031118',
                'https://cloud.cn2030.com/sc/wx/HandlerSubscribe.ashx?act=CustomerProduct&id=4847&lat=30.27415&lng=120.15515&key302=fc508a9b9c&expire302=1623031122',
                ]
    text_all = get_time() + '\n'

    k = 0

    for url in url_list:
        # print(url)
        j = 1
        while j <= 20:
            try:
                text = get_info(url,cookie,k,ip)
                # print(type(text))
                k = int(text[1])
                print(text)
                # print(text[1])
                text = text[0] + '\n' + '*'*20 +'\n'
                text_all = text_all + text
                time.sleep(2)
                # print(text)

                k = k-2
                break

            except:
                print(f'查询出错了{j}次')
                if j == 5:
                    delate_ip(ip)
                    ip = get_excel()
                    print(f'出错5次，换新IP：{ip}')
                    j += 1
                elif j ==10:
                    delate_ip(ip)
                    ip = get_excel()
                    print(f'出错10次，换新IP：{ip}')
                    j += 1

                elif j ==15:
                    delate_ip(ip)
                    ip = get_excel()
                    print(f'出错15次，换新IP：{ip}')
                    j += 1
                elif j == 20:
                    delate_ip(ip)
                    print('出错20次，查询下一家医院')

                else:
                    j +=1
                    time.sleep(2)
                # send_wechat(f'{get_time()} hpv疫苗查询出错')
                # send_telegram(f'{get_time()} hpv疫苗查询出错')
                # time.sleep(10)
        k += 2
        print(k)


    if k == 32:
        print('暂时都不可预约')
        send_wechat('九价疫苗',f'{get_time()} hpv疫苗暂时都不可预约','URL')
        send_telegram(f'{get_time()}  hpv疫苗暂时都不可预约')
    else:
        print('有疫苗可预约')
        send_wechat('九价疫苗',f'{get_time()} hpv疫苗有可以预约的了，快去准备预约','URL')
        send_telegram(f'{get_time()}  hpv疫苗有可以预约的了，快去准备预约')


    # print(text_all)
    # send_wechat(text_all)
    # send_telegram(text_all)


main()

