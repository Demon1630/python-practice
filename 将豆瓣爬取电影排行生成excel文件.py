# -*- coding: UTF-8 ...
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
import openpyxl

import requests
import re
from bs4 import BeautifulSoup
from fake_useragent import UserAgent


headers = {
    'User-Agent':UserAgent().random
}
dic_all = []
def get_content(url,dic):
# url = 'https://movie.douban.com/top250'

    respons = requests.get(url=url,headers=headers)

    # print(respons.text)
    try:
        soup = BeautifulSoup(respons.content,'html.parser')

        for j in  soup.find_all('div',attrs={'class':'item'}):    #返回的i是bs4格式，同样还可以用标签继续提取

            dic_l = []
            list = j.find('div',attrs={'class':'pic'})
            # print(type(list))
            # print(f'排名：{list.em.text}',end='\t')
            dic_l.append(list.em.text)

            for i in j.find_all('div',attrs={'class':'info'}):
                # print(f'电影名：{i.a.span}')    #继续使用a标签提取


                    #获取电影名
                    name = i.a.span.text
                    # print(f'电影名：{name}\t',end='\t\t')    #继续使用a标签提取电影名
                    dic_l.append(name)

                    #提取导演演员信息
                    info = i.find("div",attrs={"class":"bd"}).p.text.replace('\n','').replace(' ','').strip()
                    # print(f'电影信息：{info}',end='\t')    #继续使用a标签提取
                    # print(type(i.find("div",attrs={"class":"bd"})))  #继续使用a标签提取
                    dic_l.append(info)

                    # 获取评分
                    n = str(i)
                    num_list = re.search('average">(.*?)</span>',n)  #使用正则表达式提取出来
                    # print(f'评分：{num_list.group(1)}',end='\t')
                    dic_l.append(num_list.group(1))

                    #获取一句话评价
                    eval = re.search('class="inq">(.*?)</span>',n)
                    # print(type(eval))
                    # print(len(eval))
                    #添加一个判断语句
                    if eval:
                        # print(f'一句话评价：{eval.group(1)}',end='\t')
                        dic_l.append(eval.group(1))

                    else:
                        # print(f'一句话评价：无',end='\t')
                        dic_l.append('无')

                    #获取详情链接
                    url_detail = i.a.attrs['href']
                    # print(f'详情页链接为：{url_detail}',end='\n')
                    dic_l.append(url_detail)
                    # print('*'*80)
                    # print(dic)
            dic.append(dic_l)
        print(dic)
        return dic
    except Exception as result:
        print(result)
        return dic
    # return dic


def geturl_list(i):
    """使用循环获取前250电影页面链接"""
    j = i*25
    url_250 = 'https://movie.douban.com/top250?start='+str(j)+'&filter='
    # print(url_250)
    return url_250

    # print(f'电影名：{i.a.span.next_sibling.next_sibling}')    #继续使用next_sibling标签提取后面的span内容
    # print(f'电影名：{i.a.find_all("span")}')    #继续使用a标签提取
# print(type(soup.find_all('li')))


def make_excel():
    """创建excel表格"""
    wb = workbook.Workbook()  # 创建Excel对象
    ws = wb.active  # 获取当前正在操作的表对象
    # 往表中写入标题行,以列表形式写入！
    # ws.append(['角色名字', '票数'])
    ws['A1'] = '排名'
    ws['B1'] = '电影名'
    ws['C1'] = '电影信息'
    ws['D1'] = '评分'
    ws['E1'] = '一句话评价'
    ws['F1'] = '详情页链接'

    # 如果有相同文件，则直接覆盖
    wb.save("C:\\Users\\Administrator\\Desktop\\电影排名.xlsx")

def write_excel(list):
    """将传入的数据写入"""
    #先读取，再写入
    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\电影排名.xlsx')
    ws = book.active  # 获取当前正在操作的表对象
    ws.append(list)
    #如果有相同文件，则直接覆盖
    book.save("C:\\Users\\Administrator\\Desktop\\电影排名.xlsx")


def main():
    make_excel()
    for i in range(0,10):
        print(f'爬取页面{i+1}内容')
        url = geturl_list(i)
        dic = []
        dic_list = []
        dic_list=get_content(url,dic)
        for i in dic_list:
            write_excel(i)
            print(i)


if __name__ == '__main__':
    main()
