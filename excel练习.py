from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
import openpyxl
from  openpyxl import load_workbook

wb = workbook.Workbook()  # 创建Excel对象
ws = wb.active  # 获取当前正在操作的表对象
# 往表中写入标题行,以列表形式写入！
# ws.append(['角色名字', '票数'])
ws['A1'] = '排名'
ws['B1'] = '电影名'
ws['C1'] = '电影名'

rows = [
        [88, 46, 57],
        [89, 38, 12],
        [23, 59, 78],
        [56, 21, 98],
        [24, 18, 43],
        [34, 15, 67]
    ]

for row in rows:
    ws.append(row)


#如果有相同文件，则直接覆盖
wb.save("C:\\Users\\Administrator\\Desktop\\练习.xlsx")


book = openpyxl.load_workbook("C:\\Users\\Administrator\\Desktop\\练习.xlsx")
wb = book.active

#读取某一列
strs = []
for l in wb['A']:
    # c = wb[f'C{l}']
    strs.append(l.value)
    print(l.value)

print(int(88) in strs)

for i in  strs:
    print(i)

#不能直接这样写，会判断错误，还是要转换为列表
if '排名' in strs:
    print('ok')
else:
    print('no')

# print(ws)