# -*- coding: UTF-8 ...
import re
import requests
from bs4 import BeautifulSoup
from  fake_useragent import UserAgent
import os
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
import openpyxl
import wordcloud
import jieba  #分词
import telegram



headers = {
    'User-Agent':UserAgent().random
}

def get_content(url):

    response = requests.get(url=url,headers=headers)

    # print(response.text)

    soup =BeautifulSoup(response.text,'html.parser')   #必须要加上content或者text

    # print(soup)
    content = soup.tbody
    # print(content)
    # print(type(content))
    lists = content.find_all('tr')
    i = 1
    list=[]
    for l in lists:

        # print(l)
        # print(type(l))
        dic_li = []
        if i != 1:   #排除掉第一个上升热点
            info_num = l.find('td', class_='td-01 ranktop').text  #获取排名
            dic_li.append(info_num)
            # print(info_num)
            info_name = l.find('td',class_ = 'td-02').a.text   #获取关键词
            info_url = 'https://s.weibo.com'+l.find('td',class_='td-02').a.attrs['href']   #获取链接
            # print(info_name)
            # print(info_url)
            dic_li.append(info_name)
            dic_li.append(info_url)

            # print(dic_li)
            list.append(dic_li)
        i += 1


    return list
    # print(list)

def make_excel():
    """创建excel表格"""
    wb = workbook.Workbook()  # 创建Excel对象
    ws = wb.active  # 获取当前正在操作的表对象
    # 往表中写入标题行,以列表形式写入！
    # ws.append(['角色名字', '票数'])
    ws['A1'] = '排名'
    ws['B1'] = '关键词'
    ws['C1'] = '链接'

    # 如果有相同文件，则直接覆盖
    wb.save("C:\\Users\\Administrator\\Desktop\\微博热点排名.xlsx")

def write_excel(list):
    """将传入的数据写入"""
    #先读取，再写入
    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\微博热点排名.xlsx')
    ws = book.active  # 获取当前正在操作的表对象

    #把第二列关键词提取出来,判断是否存在
    strs = []
    for key_word in ws['B']:
        strs.append(key_word.value)

    if list[1] not in strs:
        print(f'{list[1]}  关键词不存在,存入并推送')
        text = list[0] + '  ' + list[1] + '  ' + list[2] + '\n'
        send_telegram(text)    #关键词之前不存在,则推送并添加进去
        ws.append(list)
        book.save("C:\\Users\\Administrator\\Desktop\\微博热点排名.xlsx")





def make_wordcloud(text):
    w = wordcloud.WordCloud(width=1000,
                            height=700,
                            background_color='white',
                            font_path='msyh.ttc')

    txtlist = jieba.lcut(text)
    string = " ".join(txtlist)

    # 将txt变量传入w的generate()方法，给词云输入文字
    w.generate(string)

    # 将词云图片导出到当前文件夹
    w.to_file('C:\\Users\\Administrator\\Desktop\\热点1.png')



def send_telegram(text):
    chat_id = '1203976293'
    token = '1865491593:AAG_GtZ7ANjE54m_9m1JGex1fQPtW9Wy8jM'
    bot = telegram.Bot(token=token)

    bot.send_message(chat_id=chat_id, text=text)


def main():
    url = 'https://s.weibo.com/top/summary?cate=realtimehot'
    # make_excel()
    try:
        list = get_content(url)
        # i = 0
        for l in list:
            if l[0].isdigit() :
                if int(l[0]) <= 20:    #获取前20的热点
                    text =  l[0] + '  ' + l[1] + '  '+ l[2] +'\n'
                    # send_telegram(text)
                    print(text)
                    # i = write_excel(l, i)
                    write_excel(l)
    except:
        send_telegram('采集出错了！')


if __name__ == '__main__':

    main()



