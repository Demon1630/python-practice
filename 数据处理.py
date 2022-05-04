import pandas
import openpyxl
import os


def chuli_excel(path):
    wb = openpyxl.load_workbook(path)
    print(wb.sheetnames)

    ws = wb['Sheet1']
    m_list = ws.merged_cells   #获取合并的单元格信息
    print(m_list)
    print(type(m_list))

#获取合并单元格内信息
    #使用split来拆分空格

    data_r=[]
    a= str(m_list).split()
    print(a)
    for i in a:
        b = i.split(':')[0]
        print(b)
        print(ws[b].value)
        data_r.append(ws[b].value)
        # print(ws['A5'].value)
    print(data_r)


    cr = []
    for m_area in m_list:
        # 合并单元格的起始行坐标、终止行坐标。。。。，
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col

        # 只对上下合并的单元格进行获取
        # if r2 - r1 > 0:
        #     cr.append((r1, r2, c1, c2))

#获取所有的合并单元格
        cr.append((r1, r2, c1, c2))
    print(cr)
        # value = ws['']

    ws.cell(15,1).value='222'

#取消单元格合并
    for r in cr:
        ws.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])

        #将取消合并后多出来的单元格内容填充原先单元格内容
        if r[1] - r[0] > 0:   #判断是不是同一行合并的多列
            a = r[1]
            b = r[0]
            while a > b:
                ws.cell(a,r[2]).value=ws.cell(b,r[2]).value  #将新单元格内容等于原先单元格内容
                # print(a,r[2])
                a -= 1
        else:
            a = r[3]
            b = r[2]
            while a > b:
                ws.cell(r[0],a).value=ws.cell(r[0],b).value
                # print(r[0],a)
                a -= 1

    #将A列数据格式设置为日期
    for cell in ws['A']:
        # cell.number_format = 'mm-dd-yy'
        cell.number_format = 'mm-dd-yy'

    wb.save(path)

def read_excel(path):

    #ExcelFile函数可以一次读取excel中的所有工作表
    xlsx = pandas.ExcelFile(path)
    sheet_nm=xlsx.sheet_names  #获取所有工作表名字
    print(sheet_nm)

    for i in sheet_nm:
        if i=="立体库出入库明细":
            data = pandas.read_excel(xlsx,i,header=1).fillna(0)
            print(data.head(3))
            print(data.iloc[1:4])
        elif i =="冷藏库出入库明细":
            data = pandas.read_excel(xlsx,i,header=1).fillna(0)
            print(data.head(3))
            print(data.iloc[1:4])
        else:
            print(i)

def get_path():
    #获取文件夹内待处理excel文件名称
    path = 'E:\python学习\零基础学python爬虫、数据分析\数据处理练习\效率需求测算'
    filelist = os.listdir(path)
    # filelist.sort()

    # filelist.sort(key=lambda x: str(x.split('内乡')[0]))
    # print(filelist)
    # for i in filelist:
    #     print(i)
    #     b = i.split('内乡')[0]
    #     print(b)

    return filelist



def main():

    filelist=get_path()
    for i in filelist:
        path = 'E:\python学习\零基础学python爬虫、数据分析\数据处理练习\效率需求测算'+i
        print(path)
    path = 'E:\牧原\牧原\项目方案\智能冷冻冷藏项目-3代厂\立体库三代方案\效率需求测算\\2021.12.1内乡厂冻品库库存V1 - 副本.xlsx'
    # read_excel(path)
    # chuli_excel(path)

if __name__ == '__main__':
    main()