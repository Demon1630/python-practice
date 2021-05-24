# -*- coding: UTF-8 ...
import re
import time

import requests
from bs4 import BeautifulSoup
from  fake_useragent import UserAgent
import os
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
import openpyxl
import wordcloud
import jieba  #分词



headers = {
    'User-Agent':UserAgent().random
}

def get_content(url):

    response = requests.get(url=url,headers=headers)

    # print(response.text)

    soup =BeautifulSoup(response.text,'html.parser')   #必须要加上content或者text

    # print(soup)
    content = soup.tbody
    # print(content)
    # print(type(content))
    lists = content.find_all('tr')
    i = 1
    list=[]
    for l in lists:

        # print(l)
        # print(type(l))
        dic_li = []
        if i != 1:   #排除掉第一个上升热点
            info_num = l.find('td', class_='td-01 ranktop').text  #获取排名
            dic_li.append(info_num)
            # print(info_num)
            info_name = l.find('td',class_ = 'td-02').a.text   #获取关键词
            info_url = 'https://s.weibo.com'+l.find('td',class_='td-02').a.attrs['href']   #获取链接
            # print(info_name)
            # print(info_url)
            dic_li.append(info_name)
            dic_li.append(info_url)

            # print(dic_li)
            list.append(dic_li)
        i += 1


    return list
    # print(list)



def make_excel():
    """创建excel表格"""
    wb = workbook.Workbook()  # 创建Excel对象
    ws = wb.active  # 获取当前正在操作的表对象
    # 往表中写入标题行,以列表形式写入！
    # ws.append(['角色名字', '票数'])
    ws['A1'] = '排名'
    ws['B1'] = '关键词'
    ws['C1'] = '链接'

    # 如果有相同文件，则直接覆盖
    wb.save("C:\\Users\\Administrator\\Desktop\\微博热点排名.xlsx")

def write_excel(list,i):
    """将传入的数据写入"""
    #先读取，再写入
    book = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\微博热点排名.xlsx')
    ws = book.active  # 获取当前正在操作的表对象

    strs = []
    for l in ws['B']:
        strs.append(l.value)
    if list[1] not in strs:
        print(f'{list[1]}  关键词不存在,存入并推送')
        ws.append(list)
        i = i
        book.save("C:\\Users\\Administrator\\Desktop\\微博热点排名.xlsx")
    else:
        i += 1
        print(f'{list[1]}  关键词存在，不存入')


    return i
    #如果有相同文件，则直接覆盖



def make_wordcloud(text):
    w = wordcloud.WordCloud(width=1000,
                            height=700,
                            background_color='white',
                            font_path='msyh.ttc')

    txtlist = jieba.lcut(text)
    string = " ".join(txtlist)

    # 将txt变量传入w的generate()方法，给词云输入文字
    w.generate(string)

    # 将词云图片导出到当前文件夹
    w.to_file('C:\\Users\\Administrator\\Desktop\\热点1.png')


def main():
    url = 'https://s.weibo.com/top/summary?cate=realtimehot'

    # print(list)
    make_excel()
    while True:
        list = get_content(url)
        print(list)
        text = ''
        i = 0
        for l in list:
            if l[0].isdigit() :
                if int(l[0]) <= 20:
                    i = write_excel(l,i)
            # print(l)
                    text += l[1]
                    text += ' '
            # i+=1
        print(i)
        if i >= 20:
            print('近一分钟无新关键词')

        # print(text)
        make_wordcloud(text)
        time.sleep(60)



if __name__ == '__main__':
    main()



